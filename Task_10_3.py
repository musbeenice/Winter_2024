# Третья задача к десятому занятию

import openpyxl

wb = openpyxl.load_workbook("test.xlsx")
ws = wb.active

lst = []

for i in range(ws.max_row):
    lst.append(ws.cell(i + 1, 2).value)
print(lst)

if "Stat" not in wb:
    wb.create_sheet("Stat")
ws = wb["Stat"]

ws.cell(1, 1).value = "Максимальное значение"
ws.cell(1, 2).value = max(lst)
ws.cell(2, 1).value = "Минимальное значение"
ws.cell(2, 2).value = min(lst)
ws.cell(3, 1).value = "Сумма"
ws.cell(3, 2).value = sum(lst)
ws.cell(4, 1).value = "Среднее значение"
ws.cell(4, 2).value = sum(lst) / len(lst) if len(lst) > 0 else 1
leng = len(lst)
if leng:
    medi = lst[leng // 2] if leng % 2 else (lst[leng // 2 - 1] + lst[leng // 2]) / 2
    ws.cell(5, 1).value = "Медиана"
    ws.cell(5, 2).value = medi

for i in range(ws.max_row):
    for j in range(ws.max_column):
        print(i + 1, j + 1, ws.cell(i + 1, j + 1).value)

wb.save("test.xlsx")
