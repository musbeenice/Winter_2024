# Вторая задача к десятому занятию

import openpyxl

wb = openpyxl.load.workbook("test.xlsx")
ws = wb.active

lst = []

for i in range(ws.max_row):
    lst.append((ws.cell(i + 1, 1).value, ws.cell(i + 1, 2).value))
print(lst)

if "New" not in wb:
    wb.create_sheet("New")
ws = wb["New"]

su = 0

for i, j in enumerate(sorted(lst, key=lambda x: (-x[i], x[0])), 1):
    ws.cell(i, 1).value = j[0]
    ws.cell(i, 2).value = j[1]
    su += j[1]

ws.cell(i + 1, 1).value = "ИТОГО"
ws.cell(i + 1, 2).value = su

for i in range(ws.max_row):
    for i in range(ws.max_column):
        print(i + 1, j + 1, ws.cell(i + 1, j + 1).value)

wb.save("test.xlsx")
