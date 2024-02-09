# Вторая задача к одиннадцатому занятию

# lst = []

# with open("test11.txt", "r", encoding="utf-8") as f:
#     for i in f:
#         lst.append(tuple(i.strip().split(", ")))
# print(lst)

# from openpyxl import *

# wb = load_workbook("test11.xlsx")
# ws = wb.active

# total = 0

# for i, j in enumerate(sorted(lst, key=lambda x: (x[3], x[1], x[2])), 1):
#     ws.cell(i, 1).value = int(j[0])
#     ws.cell(i, 2).value = j[1]
#     ws.cell(i, 3).value = j[2]
#     ws.cell(i, 4).value = j[3]
#     ws.cell(i, 5).value = int(j[4])
#     total += int(j[4])

# ws.cell(i + 1, 1).value = "ИТОГО"
# ws.cell(i + 1, 5).value = total

# wb.save("test11.xlsx")

# or

import openpyxl

with open("test11.txt", "r", encoding="utf-8") as f:
    lst = []
    for i in f:
        lst.append(i.strip().split(", "))

wb = openpyxl.load_workbook("test11.xlsx")
if "Sheet" not in wb.sheetnames:
    wb.create_sheet("Sheet")
ws = wb["Sheet"]

for i in range(1, ws.max_row):
    for j in range(1, ws.max_column):
        print(ws.cell(i, j).value == None)

lst_sorted = sorted(lst, key=lambda x: (x[3], x[1], x[2]))

total = 0
for i, j in enumerate(lst_sorted, 1):
    for k in range(5):
        if j[k].isdigit():
            ws.cell(i, k + 1).value = int(j[k])
        else:
            ws.cell(i, k + 1).value = j[k]
    total += int(j[k])

ws.cell(i + 1, 2).value = "ИТОГО"
ws.cell(i + 1, 5).value = total

wb.save("test11.xlsx")
